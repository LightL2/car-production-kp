# -*- coding: utf-8 -*-
import json
import re
from pathlib import Path

import openpyxl

PATH = Path(r"e:\Документы\Копия Смета 8Bit x BYD.xlsx")
OUT = Path(__file__).parent / "estimate_data.json"

BLOCKS = [
    ("tvc", 0, "TVC · Главный ролик", 2),
    ("photo", 5, "Photo Unit · Фотосъёмка", 3),
    ("overview", 10, "Overview · Обзорные ролики", 2),
]

RATE = 12100


def parse_uzs(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("\u00a0", " ").replace("сум", "").strip()
    s = re.sub(r"[^\d.]", "", s.replace(",", "."))
    return float(s or 0)


def fmt_usd(n):
    return f"${n:,.0f}".replace(",", " ")


def fmt_uzs(n):
    if isinstance(n, str):
        n = float(re.sub(r"[^\d.]", "", n.replace(",", ".")) or 0)
    return f"{n:,.0f}".replace(",", " ")


def parse_block(ws, start_col):
    lines = []
    section = ""
    subgroup = ""
    subtotal = None
    markup = None
    tax = None
    overall_usd = None
    overall_uzs = None

    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, start_col + 1).value
        qty = ws.cell(r, start_col + 2).value
        rate = ws.cell(r, start_col + 3).value
        amount = ws.cell(r, start_col + 4).value

        if label is None:
            continue
        text = str(label).strip()
        if not text:
            continue

        if re.match(r"^\d+\.\s", text) and qty is None and amount is None:
            section = text
            subgroup = ""
            continue

        if text == "Статья расходов":
            continue

        if text.startswith("Общая сумма"):
            if isinstance(amount, (int, float)):
                subtotal = float(amount)
            continue

        if text == "Mark up:":
            markup = float(ws.cell(r, start_col + 3).value or 0)
            continue

        if text == "Tax:":
            tax = float(ws.cell(r, start_col + 3).value or 0)
            continue

        if text == "Overall:":
            overall_usd = float(ws.cell(r, start_col + 2).value or 0)
            overall_uzs = parse_uzs(ws.cell(r, start_col + 3).value)
            continue

        if re.search(r"\$\s*[\d\s]+$", text) and qty is None:
            subgroup = re.sub(r"\s*-\s*\$.*$", "", text).strip()
            continue

        if isinstance(amount, (int, float)) and text not in ("Mark up:", "Tax:", "Overall:"):
            lines.append(
                {
                    "section": section,
                    "subgroup": subgroup,
                    "name": text,
                    "qty": qty,
                    "rate": rate,
                    "amount": float(amount),
                }
            )

    return {
        "lines": lines,
        "subtotal": subtotal,
        "markup": markup,
        "tax": tax,
        "overall_usd": overall_usd,
        "overall_uzs": overall_uzs,
    }


def rollup(lines):
    """Group line items by section for simplified table."""
    sections = []
    current = None
    for line in lines:
        sec = line["section"] or "Прочее"
        if current is None or current["section"] != sec:
            current = {"section": sec, "amount": 0, "items": []}
            sections.append(current)
        current["amount"] += line["amount"]
        current["items"].append(line)
    return sections


def main():
    wb = openpyxl.load_workbook(PATH, data_only=True)
    ws = wb.active
    data = {}
    for key, col, title, shifts in BLOCKS:
        block = parse_block(ws, col)
        block["title"] = title
        block["shifts"] = shifts
        block["sections"] = rollup(block["lines"])
        block["overall_uzs"] = parse_uzs(block["overall_uzs"])
        data[key] = block

    photo = data["photo"]
    video_tvc = data["tvc"]
    video_ov = data["overview"]

    photo_total = photo["overall_usd"]
    video_total = video_tvc["overall_usd"] + video_ov["overall_usd"]

    combined = {
        "rate": RATE,
        "photo": photo,
        "video_tvc": video_tvc,
        "video_overview": video_ov,
        "photo_total_usd": photo_total,
        "video_total_usd": video_total,
        "grand_total_usd": photo_total + video_total,
        "photo_total_uzs": parse_uzs(photo["overall_uzs"]),
        "video_total_uzs": parse_uzs(video_tvc["overall_uzs"]) + parse_uzs(video_ov["overall_uzs"]),
    }
    combined["grand_total_uzs"] = combined["photo_total_uzs"] + combined["video_total_uzs"]

    OUT.write_text(json.dumps(combined, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Photo: {fmt_usd(photo_total)} / {fmt_uzs(photo['overall_uzs'])} UZS")
    print(f"Video: {fmt_usd(video_total)} / {fmt_uzs(combined['video_total_uzs'])} UZS")
    print(f"Total: {fmt_usd(combined['grand_total_usd'])} / {fmt_uzs(combined['grand_total_uzs'])} UZS")
    print(f"Saved {OUT}")


if __name__ == "__main__":
    main()
